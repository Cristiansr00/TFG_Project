import csv
import os
from pathlib import Path
from ultralytics import YOLO
import utils.paths as paths
import numpy as np
import cv2
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
RUTA_METRICAS = 'tests'

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff"}
CLASS_COLORS = {
    "NoDicentrico": (45, 45, 220),
    "Dicentrico": (70, 180, 70),
}


def get_model_weight_files(tmodel_name):
    """
    Devuelve los pesos best.pt disponibles dentro de un modelo entrenado.
    """
    model_dir = Path(paths.MODELS_DIR) / tmodel_name
    if not model_dir.exists():
        raise FileNotFoundError(f"No existe el modelo entrenado: {model_dir}")

    weights = sorted(
        model_dir.rglob("best.pt"),
        key=lambda path: str(path).lower(),
    )
    if not weights:
        raise FileNotFoundError(f"No se encontraron pesos best.pt en {model_dir}")

    return weights


def _build_default_inference_output(image_path, model_name):
    output_dir = Path(paths.TESTS_DIR) / "inference" / model_name
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir / f"{image_path.stem}_inference.jpg"


def _class_color(class_name):
    return CLASS_COLORS.get(class_name, (220, 170, 40))


def _draw_label(image, text, x1, y1, color):
    font = cv2.FONT_HERSHEY_SIMPLEX
    font_scale = 0.55
    thickness = 1
    padding = 4
    (text_width, text_height), baseline = cv2.getTextSize(text, font, font_scale, thickness)
    label_y1 = max(0, y1 - text_height - baseline - padding * 2)
    label_y2 = min(image.shape[0], label_y1 + text_height + baseline + padding * 2)
    label_x2 = min(image.shape[1], x1 + text_width + padding * 2)

    cv2.rectangle(image, (x1, label_y1), (label_x2, label_y2), color, -1)
    cv2.putText(
        image,
        text,
        (x1 + padding, label_y2 - baseline - padding),
        font,
        font_scale,
        (255, 255, 255),
        thickness,
        cv2.LINE_AA,
    )


def run_image_inference(tmodel_name, weights_path, image_path, output_path=None, confidence=0.25):
    """
    Aplica inferencia a una imagen y guarda una copia con cajas, clase y probabilidad.
    """
    image_path = Path(image_path)
    weights_path = Path(weights_path)

    if not image_path.exists():
        raise FileNotFoundError(f"No existe la imagen: {image_path}")
    if image_path.suffix.lower() not in IMAGE_EXTENSIONS:
        raise ValueError(f"Formato de imagen no soportado: {image_path.suffix}")
    if not weights_path.exists():
        raise FileNotFoundError(f"No existen los pesos: {weights_path}")

    output_path = Path(output_path) if output_path else _build_default_inference_output(image_path, tmodel_name)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    image = cv2.imread(str(image_path))
    if image is None:
        raise ValueError(f"No se pudo leer la imagen: {image_path}")

    model = YOLO(str(weights_path))
    results = model.predict(source=str(image_path), conf=float(confidence), device="cpu", verbose=False)
    result = results[0]
    names = result.names or getattr(model, "names", {})
    annotated = image.copy()
    detections = []

    for box in result.boxes:
        class_id = int(box.cls[0])
        probability = float(box.conf[0])
        class_name = names.get(class_id, str(class_id)) if isinstance(names, dict) else str(class_id)
        x1, y1, x2, y2 = [int(round(value)) for value in box.xyxy[0].tolist()]
        x1 = max(0, min(x1, annotated.shape[1] - 1))
        y1 = max(0, min(y1, annotated.shape[0] - 1))
        x2 = max(0, min(x2, annotated.shape[1] - 1))
        y2 = max(0, min(y2, annotated.shape[0] - 1))

        color = _class_color(class_name)
        label = f"{probability:.2f}"
        cv2.rectangle(annotated, (x1, y1), (x2, y2), color, 2)
        _draw_label(annotated, label, x1, y1, color)
        detections.append({
            "class": class_name,
            "confidence": probability,
            "box": [x1, y1, x2, y2],
        })

    if not cv2.imwrite(str(output_path), annotated):
        raise OSError(f"No se pudo guardar la imagen anotada: {output_path}")

    return {
        "output_path": output_path,
        "detections": detections,
    }

def obtener_split_por_indice(splits_dir, i):
    splits_path = Path(splits_dir)
    patron = f"split_{i}"

    for d in splits_path.iterdir():
        if d.is_dir() and patron in d.name:
            return d

    raise FileNotFoundError(f"No se encontro el split {i} en {splits_dir}")


def compare_consolidated_results(excel_file=None, metric="F1-Score", metric_type=None, limit=10):
    """
    Lee el Excel consolidado y devuelve los mejores resultados por métrica.
    """
    excel_file = excel_file or f"{paths.TESTS_DIR}/consolidated_results.xlsx"
    excel_path = Path(excel_file)
    if not excel_path.exists():
        raise FileNotFoundError(f"No existe el archivo consolidado: {excel_path}")

    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    if metric not in headers:
        raise ValueError(f"La métrica '{metric}' no existe en {excel_path}")

    metric_col = headers.index(metric)
    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        if metric_type and row[1] != metric_type:
            continue
        value = row[metric_col]
        if not isinstance(value, (int, float)):
            continue
        results.append({
            "configuration": row[0],
            "metric_type": row[1],
            "value": value,
        })

    results.sort(key=lambda item: item["value"], reverse=True)
    return results[:limit]


def save_to_excel(excel_file, configuration, tmodel_name, dataset_name, metrics_data):
    """
    Guarda o actualiza los resultados de prueba en un archivo Excel consolidado.
    
    Args:
        excel_file (str): Ruta del archivo Excel
        configuration (str): Configuración del modelo
        tmodel_name (str): Nombre del modelo entrenado
        dataset_name (str): Nombre del dataset
        metrics_data (dict): Diccionario con todas las métricas
    """
    excel_path = Path(excel_file)
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Crear o cargar workbook
    if excel_path.exists():
        wb = load_workbook(excel_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Results"
        
        # Crear cabecera
        headers = [
            "Configuration", "Metric Type", "TP", "TN", "FP", "FN",
            "Recall", "Recall_Std", "Precision", "Precision_Std", "F1-Score", "F1-Score_Std",
            "SpatialAccuracy", "SpatialAccuracy_Std", "Accuracy", "Accuracy_Std"
        ]
        ws.append(headers)
        
        # Formatear cabecera
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Si la configuración ya existe, se actualiza en vez de duplicarla.
    config_name = configuration if configuration else tmodel_name
    detection_row_index = None
    classification_row_index = None
    
    for row in range(2, ws.max_row + 1):
        if ws[f'A{row}'].value == config_name:
            if ws[f'B{row}'].value == "Detection":
                detection_row_index = row
            elif ws[f'B{row}'].value == "Classification":
                classification_row_index = row
    
    # Preparar datos para escritura
    # Detection metrics
    detection_data = [
        config_name, "Detection",
        metrics_data['TP_d'], "-", metrics_data['FP_d'], metrics_data['FN_d'],
        metrics_data['recall_d'], metrics_data['recall_d_std'],
        metrics_data['precision_d'], metrics_data['precision_d_std'],
        metrics_data['fmeasure_d'], metrics_data['fmeasure_d_std'],
        metrics_data['spatialacc_d'], metrics_data['spatialacc_d_std'],
        "-", "-"
    ]
    
    # Escribir o actualizar fila de Detection
    detection_row = detection_row_index if detection_row_index else ws.max_row + 1
    for col, value in enumerate(detection_data, 1):
        ws.cell(row=detection_row, column=col, value=value)
    
    # Aplicar color de fondo a Detection (azul discreto)
    detection_fill = PatternFill(start_color="E8EFF7", end_color="E8EFF7", fill_type="solid")
    for col in range(1, len(detection_data) + 1):
        ws.cell(row=detection_row, column=col).fill = detection_fill
    
    # Classification metrics
    classification_data = [
        config_name, "Classification",
        metrics_data['TP_c'], metrics_data['TN_c'], metrics_data['FP_c'], metrics_data['FN_c'],
        metrics_data['recall_c'], metrics_data['recall_c_std'],
        metrics_data['precision_c'], metrics_data['precision_c_std'],
        metrics_data['fmeasure_c'], metrics_data['fmeasure_c_std'],
        metrics_data['spatialacc_c'], metrics_data['spatialacc_c_std'],
        metrics_data['accuracy_c'], metrics_data['accuracy_c_std']
    ]
    
    # Actualizar o insertar fila de Classification
    if classification_row_index:
        for col, value in enumerate(classification_data, 1):
            ws.cell(row=classification_row_index, column=col, value=value)
        classification_row = classification_row_index
    else:
        classification_row = ws.max_row + 1
        for col, value in enumerate(classification_data, 1):
            ws.cell(row=classification_row, column=col, value=value)
    
    # Aplicar color de fondo a Classification (verde discreto) y borde inferior
    classification_fill = PatternFill(start_color="EFF7E8", end_color="EFF7E8", fill_type="solid")
    bottom_border = Border(bottom=Side(style='medium', color="999999"))
    for col in range(1, len(classification_data) + 1):
        cell = ws.cell(row=classification_row, column=col)
        cell.fill = classification_fill
        cell.border = bottom_border
    
    # Formatear columnas numéricas
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column >= 3:  # Columnas numéricas
                if isinstance(cell.value, (int, float)):
                    # TP, TN, FP, FN (columnas 3, 4, 5, 6) con 1 decimal
                    if cell.column <= 6:
                        cell.number_format = '0.0'
                    # Resto de métricas con 4 decimales
                    else:
                        cell.number_format = '0.0000'
                cell.alignment = Alignment(horizontal="center")
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(excel_file)
    print(f"Resultados guardados en Excel: {excel_file}")


def testAndsave(configuration, tmodel_name, dataset_name, nFolds):
    """
    Esta función evalúa un modelo YOLOv8 entrenado en un conjunto de datos específico y guarda las métricas de rendimiento en un archivo CSV.
    Args:
        configuration (str): Configuración del modelo, puede ser una cadena vacía para la configuración `basic`.
        trainPath (str): Ruta al directorio del modelo YOLO ya entrenado.
        sD (str): Nombre del conjunto de datos utilizado para el entrenamiento.
        nFolds (int): Número de pliegues para la validación cruzada k-fold.
    Returns:
        None
    """

    # Fichero CSV para almacenar los resultados
    csv_file = f"{paths.TESTS_DIR}/{tmodel_name}/{dataset_name}/metrics_result.csv"
    csv_file_path = Path(paths.TESTS_DIR) / tmodel_name / dataset_name / "metrics_result.csv"
    
    # Comprueba si el archivo CSV ya existe y si la configuración ya está presente (no es necesario realmente)
    config_name = configuration if configuration else tmodel_name

    if os.path.exists(csv_file):
        with open(csv_file, mode='r') as file:
            reader = csv.reader(file)
            for row in reader:
                if row and row[0] == config_name:
                    print(f"La configuración '{config_name}' ya existe. Se omite la evaluación.")
                    return
    else:
        csv_file_path.parent.mkdir(parents=True, exist_ok=True)
    
    TP_d = 0
    FP_d = 0
    FN_d = 0
        
    TP_c = 0
    TN_c = 0
    FP_c = 0
    FN_c = 0
    
    # Guardamos valores por fold para calcular medias y desviaciones típicas.
    TP_d_values = []
    FP_d_values = []
    FN_d_values = []
    TP_c_values = []
    TN_c_values = []
    FP_c_values = []
    FN_c_values = []
    
    recall_d_values = []
    precision_d_values = []
    fmeasure_d_values = []
    spatialacc_d_values = []
    
    recall_c_values = []
    precision_c_values = []
    fmeasure_c_values = []
    spatialacc_c_values = []
    accuracy_c_values = []
    
    for j in range(1, nFolds + 1):
        # Actualiza la ruta del archivo data.yaml según el número de pliegue
        
        data_yaml = f"{paths.DATA_DIR}/processed/{dataset_name}/split_{j}/data.yaml"
        
        modifiedPath = obtener_split_por_indice(f"{paths.MODELS_DIR}/{tmodel_name}",j)
        # modifiedPath = tmodel_name.replace("split_", f"split_{j}")
        
        # Carga el modelo YOLOv8 entrenado para el pliegue actual
        model = YOLO(
            model=f"{paths.MODELS_DIR}/{tmodel_name}/{modifiedPath.name}/train/weights/best.pt"
        )
         
        # Ejecutar la validación en el conjunto de prueba y recuperar la matriz de confusión
        results = model.val(
            data=data_yaml,
            device="cpu",
            split="test"
        )
        confMatrix = results.confusion_matrix.matrix
        print(f"Matriz de confusión para el fold {j}:\n{confMatrix}")
        
        # La matriz incluye clases y fondo; se separan métricas de detección y clasificación.
        tp_d_fold = confMatrix[0][0] + confMatrix[0][1] + confMatrix[1][0] + confMatrix[1][1]
        fp_d_fold = confMatrix[0][2] + confMatrix[1][2]
        fn_d_fold = confMatrix[2][0] + confMatrix[2][1]
        
        tp_c_fold = confMatrix[1][1]
        tn_c_fold = confMatrix[0][0]
        fp_c_fold = confMatrix[1][0]
        fn_c_fold = confMatrix[0][1]
        
        # Guardar valores del fold
        TP_d_values.append(tp_d_fold)
        FP_d_values.append(fp_d_fold)
        FN_d_values.append(fn_d_fold)
        TP_c_values.append(tp_c_fold)
        TN_c_values.append(tn_c_fold)
        FP_c_values.append(fp_c_fold)
        FN_c_values.append(fn_c_fold)
        
        # Calcular métricas para este fold
        recall_d_fold = tp_d_fold / (tp_d_fold + fn_d_fold) if (tp_d_fold + fn_d_fold) > 0 else 0
        precision_d_fold = tp_d_fold / (tp_d_fold + fp_d_fold) if (tp_d_fold + fp_d_fold) > 0 else 0
        fmeasure_d_fold = 2 * (precision_d_fold * recall_d_fold) / (precision_d_fold + recall_d_fold) if (precision_d_fold + recall_d_fold) > 0 else 0
        spatialacc_d_fold = tp_d_fold / (tp_d_fold + fn_d_fold + fp_d_fold) if (tp_d_fold + fn_d_fold + fp_d_fold) > 0 else 0
        
        recall_c_fold = tp_c_fold / (tp_c_fold + fn_c_fold) if (tp_c_fold + fn_c_fold) > 0 else 0
        precision_c_fold = tp_c_fold / (tp_c_fold + fp_c_fold) if (tp_c_fold + fp_c_fold) > 0 else 0
        fmeasure_c_fold = 2 * (precision_c_fold * recall_c_fold) / (precision_c_fold + recall_c_fold) if (precision_c_fold + recall_c_fold) > 0 else 0
        spatialacc_c_fold = tp_c_fold / (tp_c_fold + fn_c_fold + fp_c_fold) if (tp_c_fold + fn_c_fold + fp_c_fold) > 0 else 0
        accuracy_c_fold = (tp_c_fold + tn_c_fold) / (tp_c_fold + tn_c_fold + fp_c_fold + fn_c_fold) if (tp_c_fold + tn_c_fold + fp_c_fold + fn_c_fold) > 0 else 0
        
        # Guardar métricas del fold
        recall_d_values.append(recall_d_fold)
        precision_d_values.append(precision_d_fold)
        fmeasure_d_values.append(fmeasure_d_fold)
        spatialacc_d_values.append(spatialacc_d_fold)
        
        recall_c_values.append(recall_c_fold)
        precision_c_values.append(precision_c_fold)
        fmeasure_c_values.append(fmeasure_c_fold)
        spatialacc_c_values.append(spatialacc_c_fold)
        accuracy_c_values.append(accuracy_c_fold)
        
        # Acumular para promedios
        TP_d += tp_d_fold
        FP_d += fp_d_fold
        FN_d += fn_d_fold
        TP_c += tp_c_fold
        TN_c += tn_c_fold
        FP_c += fp_c_fold
        FN_c += fn_c_fold
    
    # Calcular promedios
    TP_d /= nFolds
    FP_d /= nFolds
    FN_d /= nFolds
    TP_c /= nFolds
    TN_c /= nFolds
    FP_c /= nFolds
    FN_c /= nFolds
    
    # Calcular las métricas para la detección
    recall_d = TP_d / (TP_d + FN_d) if (TP_d + FN_d) > 0 else 0
    precision_d = TP_d / (TP_d + FP_d) if (TP_d + FP_d) > 0 else 0
    fmeasure_d = 2 * (precision_d * recall_d) / (precision_d + recall_d) if (precision_d + recall_d) > 0 else 0
    spatialacc_d = TP_d / (TP_d + FN_d + FP_d) if (TP_d + FN_d + FP_d) > 0 else 0
        
    # Calcular las métricas para la clasificación
    recall_c = TP_c / (TP_c + FN_c) if (TP_c + FN_c) > 0 else 0
    precision_c = TP_c / (TP_c + FP_c) if (TP_c + FP_c) > 0 else 0
    fmeasure_c = 2 * (precision_c * recall_c) / (precision_c + recall_c) if (precision_c + recall_c) > 0 else 0
    spatialacc_c = TP_c / (TP_c + FN_c + FP_c) if (TP_c + FN_c + FP_c) > 0 else 0
    accuracy_c = (TP_c + TN_c) / (TP_c + TN_c + FP_c + FN_c) if (TP_c + TN_c + FP_c + FN_c) > 0 else 0
    
    # Calcular desviaciones típicas para las métricas de detección
    recall_d_std = np.std(recall_d_values) if recall_d_values else 0
    precision_d_std = np.std(precision_d_values) if precision_d_values else 0
    fmeasure_d_std = np.std(fmeasure_d_values) if fmeasure_d_values else 0
    spatialacc_d_std = np.std(spatialacc_d_values) if spatialacc_d_values else 0
    
    # Calcular desviaciones típicas para las métricas de clasificación
    recall_c_std = np.std(recall_c_values) if recall_c_values else 0
    precision_c_std = np.std(precision_c_values) if precision_c_values else 0
    fmeasure_c_std = np.std(fmeasure_c_values) if fmeasure_c_values else 0
    spatialacc_c_std = np.std(spatialacc_c_values) if spatialacc_c_values else 0
    accuracy_c_std = np.std(accuracy_c_values) if accuracy_c_values else 0
    
    # Abrir el fichero en modo append para añadir nuevos resultados
    with open(csv_file, mode='a', newline='') as file:
        writer = csv.writer(file)
        # Escribe la cabecera si el fichero está vacío
        if file.tell() == 0:
            writer.writerow(["Configuration", "Metric Type", "TP", "TN", "FP", "FN", "Recall", "Recall_Std", "Precision", "Precision_Std", "F1-Score", "F1-Score_Std", "SpatialAccuracy", "SpatialAccuracy_Std", "Accuracy", "Accuracy_Std"])
        
        # Escribe los resultados de las métricas en el fichero CSV
        writer.writerow([configuration if configuration else f"{tmodel_name}", "Detection", float(TP_d), "-", float(FP_d), float(FN_d), recall_d, recall_d_std, precision_d, precision_d_std, fmeasure_d, fmeasure_d_std, spatialacc_d, spatialacc_d_std, "-", "-"])
        writer.writerow([configuration if configuration else f"{tmodel_name}", "Classification", float(TP_c), float(TN_c), float(FP_c), float(FN_c), recall_c, recall_c_std, precision_c, precision_c_std, fmeasure_c, fmeasure_c_std, spatialacc_c, spatialacc_c_std, accuracy_c, accuracy_c_std])
    
    # Preparar datos para Excel
    metrics_data = {
        'TP_d': float(TP_d), 'FP_d': float(FP_d), 'FN_d': float(FN_d),
        'recall_d': recall_d, 'recall_d_std': recall_d_std,
        'precision_d': precision_d, 'precision_d_std': precision_d_std,
        'fmeasure_d': fmeasure_d, 'fmeasure_d_std': fmeasure_d_std,
        'spatialacc_d': spatialacc_d, 'spatialacc_d_std': spatialacc_d_std,
        'TP_c': float(TP_c), 'TN_c': float(TN_c), 'FP_c': float(FP_c), 'FN_c': float(FN_c),
        'recall_c': recall_c, 'recall_c_std': recall_c_std,
        'precision_c': precision_c, 'precision_c_std': precision_c_std,
        'fmeasure_c': fmeasure_c, 'fmeasure_c_std': fmeasure_c_std,
        'spatialacc_c': spatialacc_c, 'spatialacc_c_std': spatialacc_c_std,
        'accuracy_c': accuracy_c, 'accuracy_c_std': accuracy_c_std
    }
    
    # Guardar en Excel consolidado
    excel_file = f"{paths.TESTS_DIR}/consolidated_results.xlsx"
    save_to_excel(excel_file, configuration, tmodel_name, dataset_name, metrics_data)
